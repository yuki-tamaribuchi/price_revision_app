import numpy as np
import pandas as pd
from openpyxl import Workbook,load_workbook

class PriceRevision():

    def load_master(self,filepath):
        self.__master_df=pd.read_csv(filepath_or_buffer=filepath,encoding='CP932',low_memory=False,header=0)

    def create_revision_excel(self,dir_path):
        revision_wb=Workbook()
        revision_ws=revision_wb.active
        revision_ws.title='改定表貼付シート'
        revision_ws['A1']='JANコード'
        revision_ws['B1']='品番'
        revision_ws['C1']='商品名'
        revision_ws['D1']='新価格'
        revision_ws['E1']='備考1'
        revision_ws['F1']='備考2'

        revision_wb.save('%s/revision.xlsx'%dir_path)

    def create_maker_list(self):
        self.makers=list(self.__master_df['メーカー名称'].unique())

    def query_by_maker(self,maker):
        self.__queried_df=self.__master_df[self.__master_df['メーカー名称']==maker]

    def create_revision_data_frame(self,dir_path):
        self.__revision_df=pd.read_excel('%s/revision.xlsx'%dir_path)

    def execute_search(self):
        self.__matched_df=self.__revision_df[self.__revision_df['JANコード'].isin(self.__queried_df['商品コード'])]


    def export_matched_data(self,dir_path):
        with pd.ExcelWriter('%s/revision.xlsx'%dir_path) as writer:
            writer.book=load_workbook('%s/revision.xlsx'%dir_path)
            self.__matched_df.to_excel(writer,sheet_name='抽出済')